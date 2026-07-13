import json
import os
from datetime import datetime

import pandas as pd

# EXCEL_FILENAME removed, generating dynamically
CONFIG_FILE = "config.json"


class DataService:
    @staticmethod
    def load_config():
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if not content:
                        return {}
                    return json.loads(content)
            except (json.JSONDecodeError, Exception) as e:
                print(f"Error loading config: {e}. Resetting to defaults.")
                try:
                    os.rename(CONFIG_FILE, CONFIG_FILE + ".corrupted")
                except Exception:
                    pass
        return {}

    @staticmethod
    def save_config(config):
        try:
            tmp_file = CONFIG_FILE + ".tmp"
            with open(tmp_file, "w", encoding="utf-8") as f:
                json.dump(config, f, indent=4)
            os.replace(tmp_file, CONFIG_FILE)
            return True
        except Exception as e:
            print(f"Error saving config: {e}")
            try:
                os.remove(CONFIG_FILE + ".tmp")
            except Exception:
                pass
            return False

    @staticmethod
    def _apply_excel_styling(workbook):
        from openpyxl.styles import Alignment, Border, Side

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

        if "Calculo" in workbook.sheetnames:
            ws_calc = workbook["Calculo"]
            for row in ws_calc.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

        if "Detalle" in workbook.sheetnames:
            ws_detail = workbook["Detalle"]
            for row in ws_detail.iter_rows(min_row=1):
                for cell in row:
                    cell.alignment = Alignment(horizontal="center")
                    cell.border = thin_border

    @staticmethod
    def get_excel_path(folder_path=None):
        config = DataService.load_config()
        base_folder = config.get("export_folder") or folder_path or os.getcwd()

        current_year = datetime.now().year
        filename = f"Peajes {current_year} Calculo.xlsx"
        return os.path.join(base_folder, filename), filename

    @staticmethod
    def save_toll_entry(folder_path, data):
        """
        Saves a toll entry to the Excel file in the specified folder.
        Creates the file if it doesn't exist.
        The filename contains the current year (e.g., "Peajes 2026 Calculo.xlsx").
        Format:
        Row 1: Calculo peajes [Year]
        Row 2: Numero de Peajes | Total en BS
        Following rows: Sequential numbering | Amount

        Args:
            folder_path (str): Default directory if no export folder is configured.
            data (dict): Dictionary containing row data (PDF Name, Page, Amount, etc.)
        """
        file_path, filename = DataService.get_excel_path(folder_path)
        current_year = datetime.now().year

        # Add timestamp
        data["Timestamp"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        # Ensure Total Amount is a number (float)
        # It might come in as a string with formatting (e.g. from the UI entry)
        try:
            val = data.get("Total Amount", 0)
            if isinstance(val, str):
                # Clean up any currency symbols or commas just in case
                val = val.replace("$", "").replace(",", "")
            data["Total Amount"] = float(val)
        except (ValueError, TypeError):
            data["Total Amount"] = 0.0

        try:
            # 1. Determine next sequence number
            next_num = 1
            if os.path.exists(file_path):
                try:
                    df_calc = pd.read_excel(file_path, sheet_name="Calculo", skiprows=1)
                    if not df_calc.empty:
                        last_num = df_calc.iloc[:, 0].max()
                        if pd.notna(last_num):
                            next_num = int(last_num) + 1
                except Exception:
                    next_num = 1

            # 2. Prepare data for Detalle sheet
            # Add sequential 'No.' as the first column
            detail_data = {"No.": next_num}
            detail_data.update(data)
            df_detail_new = pd.DataFrame([detail_data])

            # 3. Prepare data for Calculo sheet
            df_summary_new = pd.DataFrame(
                {
                    "Numero de Peajes": [next_num],
                    "Total en BS": [data.get("Total Amount", 0)],
                }
            )

            # 4. Write to Excel
            if os.path.exists(file_path):
                with pd.ExcelWriter(
                    file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay"
                ) as writer:
                    # Write Summary (Calculo)
                    start_row_summary = 0
                    try:
                        writer.book["Calculo"]
                        start_row_summary = writer.book["Calculo"].max_row
                        df_summary_new.to_excel(
                            writer,
                            sheet_name="Calculo",
                            index=False,
                            header=False,
                            startrow=start_row_summary,
                        )
                    except KeyError:
                        # Create sheet if it somehow disappeared but file exists
                        title_df = pd.DataFrame([[f"Calculo peajes {current_year}"]])
                        title_df.to_excel(
                            writer, sheet_name="Calculo", index=False, header=False
                        )
                        df_summary_new.to_excel(
                            writer,
                            sheet_name="Calculo",
                            index=False,
                            header=True,
                            startrow=1,
                        )

                    # Write Detalle (Detalle)
                    start_row_detail = 0
                    try:
                        writer.book["Detalle"]
                        start_row_detail = writer.book["Detalle"].max_row
                        df_detail_new.to_excel(
                            writer,
                            sheet_name="Detalle",
                            index=False,
                            header=False,
                            startrow=start_row_detail,
                        )
                    except KeyError:
                        df_detail_new.to_excel(
                            writer, sheet_name="Detalle", index=False, header=True
                        )

                    # Apply styling (Centering and Borders)
                    DataService._apply_excel_styling(writer.book)
            else:
                # Create new file
                with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                    # Title row
                    title_df = pd.DataFrame([[f"Calculo peajes {current_year}"]])
                    title_df.to_excel(
                        writer, sheet_name="Calculo", index=False, header=False
                    )

                    # Summary
                    df_summary_new.to_excel(
                        writer,
                        sheet_name="Calculo",
                        index=False,
                        header=True,
                        startrow=1,
                    )

                    # Detail
                    df_detail_new.to_excel(
                        writer, sheet_name="Detalle", index=False, header=True
                    )

                    # Apply styling (Centering and Borders)
                    DataService._apply_excel_styling(writer.book)

            return True, f"Saved to {filename}"

        except Exception as e:
            import traceback

            traceback.print_exc()
            print(f"Error saving to Excel: {e}")
            return False, str(e)

    @staticmethod
    def has_toll_entry(pdf_name, page_number, folder_path=None):
        """
        Returns True if an entry exists in the Detalle sheet for the given PDF + page.
        """
        file_path, _ = DataService.get_excel_path(folder_path)
        if not os.path.exists(file_path):
            return False
        try:
            df = pd.read_excel(file_path, sheet_name="Detalle")
            mask = (df["PDF Name"].astype(str) == str(pdf_name)) & \
                   (df["Page Number"].astype(int) == int(page_number))
            return mask.any()
        except Exception:
            return False

    @staticmethod
    def delete_toll_entry(pdf_name, page_number, folder_path=None):
        """
        Deletes a toll entry from the Excel file by matching PDF Name and Page Number.
        Removes the corresponding row from both 'Detalle' and 'Calculo' sheets.
        Does NOT renumber remaining entries (toll numbers map to physical paper).

        Returns:
            (bool, str): Success flag and message.
        """

        file_path, filename = DataService.get_excel_path(folder_path)

        if not os.path.exists(file_path):
            return False, "Excel file not found."

        try:
            # Read Detalle to find the matching row
            df_detail = pd.read_excel(file_path, sheet_name="Detalle")

            # Find matching row(s)
            mask = (df_detail["PDF Name"].astype(str) == str(pdf_name)) & \
                   (df_detail["Page Number"].astype(int) == int(page_number))

            matches = df_detail[mask]
            if matches.empty:
                return False, f"No entry found for '{pdf_name}' page {page_number}."

            # Get the No. value(s) to delete from Calculo
            entry_numbers = matches["No."].tolist()

            # Remove from Detalle
            df_detail = df_detail[~mask]

            # Read and filter Calculo
            df_calculo = pd.read_excel(file_path, sheet_name="Calculo", skiprows=1)
            calculo_mask = df_calculo["Numero de Peajes"].isin(entry_numbers)
            df_calculo = df_calculo[~calculo_mask]

            # Rewrite the Excel file
            current_year = datetime.now().year

            with pd.ExcelWriter(file_path, engine="openpyxl") as writer:
                # Write Calculo sheet
                title_df = pd.DataFrame([[f"Calculo peajes {current_year}"]])
                title_df.to_excel(
                    writer, sheet_name="Calculo", index=False, header=False
                )
                if not df_calculo.empty:
                    df_calculo.to_excel(
                        writer, sheet_name="Calculo", index=False, header=True, startrow=1
                    )
                else:
                    # Write just headers if empty
                    pd.DataFrame(columns=["Numero de Peajes", "Total en BS"]).to_excel(
                        writer, sheet_name="Calculo", index=False, header=True, startrow=1
                    )

                # Write Detalle sheet
                df_detail.to_excel(
                    writer, sheet_name="Detalle", index=False, header=True
                )

                # Apply styling
                DataService._apply_excel_styling(writer.book)

            return True, f"Removed toll #{entry_numbers[0]} from {filename}"

        except Exception as e:
            import traceback
            traceback.print_exc()
            return False, str(e)

    @staticmethod
    def get_processed_tolls(folder_path=None):
        """
        Returns a set of PDF filenames that have already been processed
        (appear in the 'PDF Name' column of the 'Detalle' sheet).
        """
        try:
            file_path, _ = DataService.get_excel_path(folder_path)
            if not os.path.exists(file_path):
                return set()

            # Read only the 'PDF Name' column from 'Detalle' sheet if it exists
            # We don't know the exact column index, but we know the header "PDF Name"
            # It's safer to read the whole sheet or specific columns by name if possible.
            # However, pd.read_excel might fail if sheet doesn't exist.
            try:
                df = pd.read_excel(file_path, sheet_name="Detalle")
                if "PDF Name" in df.columns:
                    # Return set of non-null values
                    return set(df["PDF Name"].dropna().astype(str).unique())
            except ValueError:
                # Sheet 'Detalle' usually raises ValueError if not found in some pandas versions/engines
                pass
            except Exception as e:
                print(f"Error reading processed tolls: {e}")

            return set()
        except Exception as e:
            print(f"Error in get_processed_tolls: {e}")
            return set()

    @staticmethod
    def load_flags():
        """
        Returns a set of file paths that are flagged for review.
        """
        flags_file = "flags.json"
        if os.path.exists(flags_file):
            try:
                with open(flags_file, "r", encoding="utf-8") as f:
                    content = f.read().strip()
                    if not content:
                        return set()
                    data = json.loads(content)
                    return set(data)
            except (json.JSONDecodeError, Exception) as e:
                print(f"Error loading flags: {e}. Resetting flags.")
                try:
                    os.rename(flags_file, flags_file + ".corrupted")
                except Exception:
                    pass
        return set()

    @staticmethod
    def save_flags(flags_set):
        """
        Saves the set of flagged file paths to flags.json.
        Uses atomic write to avoid corruption.
        """
        flags_file = "flags.json"
        tmp_file = flags_file + ".tmp"
        try:
            with open(tmp_file, "w", encoding="utf-8") as f:
                json.dump(list(flags_set), f, indent=4)
            os.replace(tmp_file, flags_file)
            return True
        except Exception as e:
            print(f"Error saving flags: {e}")
            try:
                os.remove(tmp_file)
            except Exception:
                pass
            return False
